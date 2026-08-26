#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
แปลง data/source/RABBiZBuild_v9_1.xlsx -> data/seed/*.json
รันครั้งเดียวก็พอ ผลลัพธ์ commit ไว้ในรีโป (server/setup.js อ่านไฟล์ JSON เหล่านี้)

  python3 tools/xlsx_to_seed.py

v9 เพิ่มจาก v8: ประเภทโครงการ · สิทธิ์รายโครงการ (กฎ 3%) · เส้นโค้งต้นทุน ·
ทะเบียนชื่อพ้องอาคาร · ทะเบียนพนักงาน · การจับคู่ผู้รับเงิน · ฝั่งรายได้ (เก็บเป็นข้อมูลอ้างอิง)
"""
import json, os, sys, unicodedata
from datetime import datetime, date
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'data', 'source', 'RABBiZBuild_v9_1.xlsx')
OUT = os.path.join(ROOT, 'data', 'seed')
YEAR_CE = 2026  # 2569 พ.ศ.

def s(v):
    if v is None: return ''
    if isinstance(v, (datetime, date)): return v.strftime('%Y-%m-%d')
    if isinstance(v, float) and v.is_integer(): return str(int(v))
    return unicodedata.normalize('NFC', str(v)).strip()

def num(v):
    if v is None or v == '': return None
    try: return round(float(v), 2)
    except (TypeError, ValueError): return None

def sheet(wb, name, header_row=1):
    """คืน list ของ dict โดยใช้แถว header_row (1-based) เป็นชื่อคอลัมน์"""
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [s(c) for c in rows[header_row - 1]]
    out = []
    for r in rows[header_row:]:
        if all(c is None or s(c) == '' for c in r): continue
        out.append({hdr[i]: r[i] for i in range(min(len(hdr), len(r))) if hdr[i]})
    return out

def grid(wb, name, header_row):
    """คืน (หัวคอลัมน์, แถวดิบ) สำหรับชีตที่เป็นตารางกากบาท"""
    rows = [[s(c) for c in r] for r in wb[name].iter_rows(values_only=True)]
    return rows[header_row - 1], rows[header_row:]

def write(name, data):
    p = os.path.join(OUT, name + '.json')
    with open(p, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f'  {name:<22} {len(data):>5} แถว')

# ---------------------------------------------------------------- ผู้ใช้
# บทบาทยึดตาม 21_DIM_USER_ACCESS ของ v9 (J และ T เปลี่ยนเป็น PM · O ลาออกแล้ว)
ROLE_BY_ID = {
    'A': 'CEO', 'K': 'COO', 'S': 'FINANCE', 'N': 'ACCOUNT',
    'M': 'PM', 'F': 'PM', 'O': 'PM', 'B': 'PM', 'J': 'PM', 'T': 'PM', 'I': 'VIEWER',
}
TITLE_BY_ID = {
    'A': 'CEO', 'K': 'COO', 'S': 'การเงิน', 'N': 'บัญชี',
    'M': 'PM', 'F': 'PM', 'O': 'PM', 'B': 'PM', 'J': 'PM', 'T': 'PM', 'I': 'ผู้ดูข้อมูล',
}
# คนที่ยังไม่มีอีเมล/เบอร์โทรในชีต 05_DIM_PERSON — สร้างไว้แต่ระงับ รอ CEO กรอกของจริง
NO_CONTACT_NOTE = 'ยังไม่มีอีเมล/เบอร์โทรในฐาน v9 — CEO ต้องกรอกและเปิดใช้งานก่อน'

def build_users(wb):
    access_names, access_roles = {}, {}
    for r in sheet(wb, '21_DIM_USER_ACCESS', header_row=4):
        uid = s(r.get('รหัส'))
        if len(uid) != 1: continue
        access_names[uid] = s(r.get('ชื่อ'))
        access_roles[uid] = s(r.get('บทบาท'))

    people = {}
    for r in sheet(wb, '05_DIM_PERSON'):
        uid = s(r.get('person_id'))
        if len(uid) == 1 and uid in ROLE_BY_ID: people[uid] = r

    out = []
    for uid, role in ROLE_BY_ID.items():
        p = people.get(uid, {})
        phone = s(p.get('เบอร์โทร (ใช้ติดต่อ)'))
        username = s(p.get('username (อีเมล)')).lower()
        resigned = 'ลาออก' in access_roles.get(uid, '')
        notes = []
        if resigned: notes.append('ลาออกแล้ว — ปิดสิทธิ์ทั้งหมด เก็บประวัติการเบิกไว้')
        if not username: notes.append(NO_CONTACT_NOTE)
        out.append({
            'user_id': uid,
            'display_name': access_names.get(uid) or s(p.get('ชื่อที่ใช้เรียก')) or uid,
            'full_name': s(p.get('ชื่อ-สกุล (ตาม ภ.ง.ด.1)')) or s(p.get('ชื่อเต็มตามที่พบในไฟล์')),
            'title': TITLE_BY_ID[uid],
            'role': role,
            'username': username or f'{uid.lower()}.pending@rabbiz.local',
            'phone': phone,
            'require_2fa': 1 if role in ('CEO', 'COO', 'FINANCE') else 0,
            'status': 'ใช้งาน' if (username and not resigned) else 'ระงับ',
            'note': ' · '.join(notes),
        })
    return out

def build_user_projects(wb, valid_projects):
    """21_DIM_USER_ACCESS — ตารางกากบาท: คอลัมน์เป็น 'RMT\\nไทยรามัญ'"""
    hdr, rows = grid(wb, '21_DIM_USER_ACCESS', 4)
    cols = {i: h.split('\n')[0].strip() for i, h in enumerate(hdr)
            if len(h.split('\n')[0].strip()) == 3 and h.split('\n')[0].strip().isupper()}
    out = []
    for r in rows:
        uid = r[0] if r else ''
        if len(uid) != 1 or uid not in ROLE_BY_ID: continue
        for i, code in cols.items():
            if i < len(r) and r[i].upper() == 'Y' and code in valid_projects:
                out.append({'user_id': uid, 'project_id': code})
    return out

# ---------------------------------------------------------------- โครงการ
def build_projects(wb):
    types = {}
    for r in sheet(wb, '29_DIM_PROJECT_TYPE', header_row=4):
        code = s(r.get('รหัส'))
        if len(code) != 3 or not code.isupper(): continue
        types[code] = r

    out, seen = [], set()
    for r in sheet(wb, '02_DIM_PROJECT'):
        pid = s(r.get('project_id'))
        if not pid or len(pid) > 4: continue
        seen.add(pid)
        t = types.get(pid, {})
        out.append({
            'project_id': pid,
            'project_name': s(t.get('ชื่อ')) or s(r.get('project_name')),
            'project_type': s(t.get('ประเภท')) or ('ค่าใช้จ่ายส่วนกลาง' if pid == 'OFF' else ''),
            'asset_status': s(t.get('สถานะ')),
            'is_group_asset': 1 if s(t.get('นับเป็นทรัพย์สินกลุ่ม')) == 'Y' else 0,
            'is_real_project': 0 if pid in ('NA', 'OFF') else 1,
            'note': s(t.get('หมายเหตุ')) or s(r.get('หมายเหตุ')),
            'status': 'ใช้งาน',
        })
    # โครงการที่มีเฉพาะใน 29 (เช่น KAN สร้างเสร็จก่อน มี.ค. จึงไม่มีรายการเบิกในชุดข้อมูล)
    for pid, t in types.items():
        if pid in seen: continue
        out.append({
            'project_id': pid,
            'project_name': s(t.get('ชื่อ')),
            'project_type': s(t.get('ประเภท')),
            'asset_status': s(t.get('สถานะ')),
            'is_group_asset': 1 if s(t.get('นับเป็นทรัพย์สินกลุ่ม')) == 'Y' else 0,
            'is_real_project': 1,
            'note': s(t.get('หมายเหตุ')),
            'status': 'ใช้งาน',
        })
    return out

# ---------------------------------------------------------------- แบบอาคาร
def build_designs(wb):
    out = []
    for r in sheet(wb, '14_DIM_DESIGN', header_row=4):
        code = s(r.get('design_code'))
        if not code.startswith('D-'): continue
        out.append({
            'design_code': code,
            'design_name': s(r.get('ชื่อแบบ')),
            'floors': int(num(r.get('จำนวนชั้น')) or 0) or None,
            'std_area_sqm': num(r.get('พื้นที่ (ตร.ม.)')),
            'structure': s(r.get('ลักษณะโครงสร้าง')),
            'ref_cost_per_sqm': num(r.get('ต้นทุนอ้างอิง/ตร.ม.')),
            'status': 'ยืนยันแล้ว' if s(r.get('สถานะ')) == 'ยืนยันแล้ว' else 'รอยืนยัน',
            'note': s(r.get('หมายเหตุ')),
        })
    return out

def build_cost_curve(wb):
    """15_DIM_COST_CURVE — ใช้ตั้งงบอาคารใหม่จากขนาดและจำนวนชั้น"""
    out, n = [], 0
    for r in sheet(wb, '15_DIM_COST_CURVE', header_row=4):
        floors, area = num(r.get('จำนวนชั้น')), num(r.get('พื้นที่ (ตร.ม.)'))
        if not floors or not area: continue
        n += 1
        out.append({
            'curve_id': f'CC{n:03d}',
            'floors': int(floors),
            'area_sqm': area,
            'design_code': s(r.get('design_code')) or None,
            'building_label': s(r.get('อาคาร')),
            'total_cost': num(r.get('ต้นทุน')),
            'cost_per_sqm': num(r.get('บาท/ตร.ม.')),
            'source': s(r.get('ที่มา')),
            'completeness': s(r.get('ความครบถ้วน')),
        })
    return out

# ---------------------------------------------------------------- อาคาร
WORK_NATURE = ('สร้างใหม่', 'ต่อเติม', 'ซ่อมบำรุง')

def build_buildings(wb):
    out = []
    for r in sheet(wb, '03_DIM_BUILDING'):
        bid = s(r.get('building_id'))
        if not (len(bid) == 4 and bid[0] == 'B' and bid[1:].isdigit()): continue
        raw_nature = s(r.get('work_nature (สร้างใหม่/ต่อเติม/ซ่อมบำรุง)'))
        nature = next((w for w in WORK_NATURE if raw_nature.startswith(w)), 'สร้างใหม่')
        status = 'ปิดจบ' if s(r.get('status (กำลังทำ/ปิดจบ)')).startswith('ปิดจบ') else 'กำลังทำ'
        extra = raw_nature[len(nature):].strip(' —-') if raw_nature.startswith(nature) else raw_nature
        note = ' · '.join(x for x in (s(r.get('หมายเหตุ')), extra) if x)
        src = s(r.get('ที่มาของค่า'))
        out.append({
            'building_id': bid,
            'project_id': s(r.get('project_id')),
            'building_name': s(r.get('building_name')),
            'design_code': s(r.get('design_code (แบบอาคาร)')) or None,
            'work_nature': nature,
            'status': status,
            'area_sqm': num(r.get('area_sqm (กรอกเมื่อเป็นงานสร้างใหม่)')),
            'floors': int(num(r.get('จำนวนชั้น')) or 0) or None,
            'is_building': 'N' if s(r.get('is_building (Y/N)')) == 'N' else 'Y',
            'budget': None,
            'value_source': 'ข้อเท็จจริง' if src.startswith('ข้อเท็จจริง')
                            else 'อนุมาน' if src.startswith('อนุมาน') else 'นำเข้าย้อนหลัง',
            'note': note,
        })
    return out

def build_building_aliases(wb):
    """20_DIM_BUILDING_ALIAS — ค้นก่อนสร้างอาคารใหม่เสมอ ไม่งั้นต้นทุนแตกเป็นสองก้อน"""
    out = []
    for r in sheet(wb, '20_DIM_BUILDING_ALIAS', header_row=4):
        bid = s(r.get('building_id'))
        if not (len(bid) == 4 and bid[0] == 'B'): continue
        for field, kind in [('ชื่อทางการ (ฝั่งรายได้)', 'ชื่อทางการฝั่งรายได้'),
                            ('ชื่อที่ใช้ในงานก่อสร้าง', 'ชื่อที่ใช้หน้างาน'),
                            ('รหัสเดิม 3 ตัว', 'รหัสเดิม')]:
            alias = s(r.get(field))
            if alias: out.append({'building_id': bid, 'alias': alias, 'alias_kind': kind})
    return out

# ---------------------------------------------------------------- หมวดงาน
GROUPS = {'1 งานโครงสร้าง': 1, '2 งานสถาปัตย์': 2, '3 งานระบบ': 3,
          '4 งานภายนอกโครงการ': 4, '5 งานอื่น/ดำเนินการ': 5}

def build_cost_codes(wb):
    out = []
    for r in sheet(wb, '04_DIM_COSTCODE'):
        code = s(r.get('cost_code'))
        if not (2 <= len(code) <= 4 and code.isupper()): continue
        status = s(r.get('สถานะ v8')) or 'ใช้ต่อ'
        if code == 'NA': status = 'เลิกใช้'   # "ไม่ระบุ" ไม่ใช่หมวดงาน — ห้ามให้เลือกในใบใหม่
        grp = s(r.get('กลุ่มงานตาม BOQ'))
        out.append({
            'cost_code': code,
            'cost_name': s(r.get('cost_name')),
            'work_group': grp if grp != '—' else '',
            'group_order': GROUPS.get(grp, 9),
            'status': status,
            'merge_into': s(r.get('ยุบเข้ารหัส')) or None,
            'default_cost_type': {'ของ': 'ของ', 'แรง': 'แรง', 'เช่า': 'เช่า', 'โสหุ้ย': 'โสหุ้ย'}.get(
                s(r.get('แกน 2 ประเภทต้นทุน')), None),
            'note': s(r.get('เหตุผล / วิธีแก้')),
        })
    return out

# ---------------------------------------------------------------- พนักงาน + ผู้ขาย
def build_employees(wb):
    out = []
    for r in sheet(wb, '18_DIM_EMPLOYEE', header_row=4):
        eid = s(r.get('emp_id'))
        if not eid.startswith('E'): continue
        out.append({
            'emp_id': eid,
            'full_name': s(r.get('ชื่อ-สกุล')),
            'employer': s(r.get('นิติบุคคลนายจ้าง')),
            'employment_type': s(r.get('ประเภทการจ้าง')),
            'base_salary': num(r.get('ค่าจ้างพื้นฐาน/เดือน')),
            'allowance': num(r.get('เบี้ยเลี้ยงประจำ')),
            'social_security': s(r.get('ประกันสังคม')),
            'start_date': s(r.get('วันเริ่มงาน')) or None,
            'end_date': s(r.get('วันสิ้นสุด')) or None,
            'status': s(r.get('สถานะ')),
            'risk_flag': s(r.get('ธงความเสี่ยง')),
        })
    return out

def build_vendors(wb):
    """06_DIM_VENDOR + 30_VENDOR_MATCH — ทำเครื่องหมายผู้รับเงินที่จริง ๆ แล้วเป็นพนักงานของเราเอง"""
    match = {}
    for r in sheet(wb, '30_VENDOR_MATCH', header_row=4):
        name = s(r.get('ผู้รับเงินในระบบ'))
        if not name: continue
        match[name] = {
            'suggested': s(r.get('ประเภทที่ควรเป็น')),
            'matched_to': s(r.get('จับคู่กับสมุดราคา')),
            'paid': num(r.get('ยอดจ่าย')),
        }

    out = []
    for r in sheet(wb, '06_DIM_VENDOR'):
        vid = s(r.get('vendor_id'))
        if not vid.startswith('V'): continue
        name = s(r.get('vendor_name'))
        vtype = s(r.get('vendor_type'))
        m = match.get(name, {})
        suggested = m.get('suggested', '')
        is_staff = 1 if 'พนักงานของเรา' in suggested else 0
        # "พนักงานของเรา — จตุพร (J)" → ดึงรหัสผู้ใช้ในวงเล็บ
        staff_user = ''
        if is_staff and '(' in suggested and ')' in suggested:
            staff_user = suggested[suggested.rfind('(') + 1:suggested.rfind(')')].strip()
        entity = ('บุคคลธรรมดา' if is_staff or 'บุคคล' in vtype or 'ช่าง' in vtype
                  or 'ผู้รับเหมา' in vtype or 'บุคคลธรรมดา' in suggested else 'นิติบุคคล')
        out.append({
            'vendor_id': vid,
            'vendor_name': name,
            'vendor_type': vtype,
            'category': s(r.get('หมวดสินค้า/บริการ')),
            'phone': s(r.get('โทรศัพท์')),
            'entity_type': entity,
            'tax_id': '', 'bank_account': '', 'payment_terms': '',
            'vat_registered': 0,
            'wht_percent': 3 if entity == 'บุคคลธรรมดา' else 0,
            'doc_status': 'ยืนยันแล้ว' if s(r.get('ยืนยันประเภทแล้ว')) else 'รอตรวจเอกสาร',
            'is_own_staff': is_staff,
            'staff_user_id': staff_user if len(staff_user) == 1 else None,
            'match_note': suggested,
            'status': 'ใช้งาน',
        })
    return out

# ---------------------------------------------------------------- วัสดุ · ราคา · rate
def build_items(wb):
    return [{
        'item_id': s(r.get('item_id')), 'category': s(r.get('category')),
        'item_name': s(r.get('item_name')), 'unit': s(r.get('unit')),
        'ref_price_min': num(r.get('ราคาต่ำสุด')), 'ref_price_max': num(r.get('ราคาสูงสุด')),
        'vendor_count': int(num(r.get('จำนวนร้านที่มีราคา')) or 0), 'status': 'ใช้งาน',
    } for r in sheet(wb, '07_DIM_ITEM') if s(r.get('item_id')).startswith('I')]

def build_item_prices(wb):
    return [{
        'price_id': s(r.get('price_id')), 'item_id': s(r.get('item_id')),
        'vendor_id': s(r.get('vendor_id')) or None, 'unit_price': num(r.get('unit_price')),
        'unit': s(r.get('unit')), 'source_note': s(r.get('source_note')),
        'is_cheapest': 1 if s(r.get('is_cheapest')) in ('1', 'True') else 0,
    } for r in sheet(wb, '11_FACT_PRICE') if s(r.get('price_id')).startswith('P')]

def build_rates(wb):
    out, n = [], 0
    for r in sheet(wb, '13_DIM_RATE', header_row=4):
        name = s(r.get('รายการ'))
        if not name: continue
        n += 1
        out.append({
            'rate_id': f'R{n:03d}',
            'cost_type': {'ค่าแรง': 'แรง', 'ค่าวัสดุ': 'ของ', 'ค่าเครื่องจักร': 'เช่า'}.get(s(r.get('ประเภท')), 'ของ'),
            'rate_name': name, 'unit': s(r.get('หน่วย')),
            'rate_satoshi': num(r.get('Satoshi L')), 'rate_goldy': num(r.get('Goldy')),
            'std_rate': num(r.get('อัตรามาตรฐาน')), 'method': s(r.get('วิธีได้มา')),
            'status': 'ยืนยันแล้ว' if s(r.get('สถานะ')) == 'ยืนยันแล้ว' else 'รอยืนยัน',
        })
    return out

# ---------------------------------------------------------------- BOQ · คู่อาคาร · เงินทุน
def build_boq(wb):
    reg = [{
        'boq_id': s(r.get('boq_id')), 'title': s(r.get('building_name')),
        'version': s(r.get('version')), 'received_date': s(r.get('วันที่รับ BOQ')),
        'source_file': s(r.get('ชื่อไฟล์ต้นทาง')),
        'boq_value': num(r.get('มูลค่า BOQ ต่อ 1 อาคาร (บาท)')),
        'author': s(r.get('ผู้จัดทำ')), 'status': s(r.get('สถานะ')),
    } for r in sheet(wb, '30_BOQ_REGISTER') if s(r.get('boq_id')).startswith('BOQ-')]

    bld = [{
        'boq_id': s(r.get('boq_id')), 'building_id': s(r.get('building_id')),
        'boq_budget': num(r.get('งบตาม BOQ (บาท)')), 'status': s(r.get('สถานะ')),
        'note': s(r.get('หมายเหตุ')),
    } for r in sheet(wb, '32_BOQ_BUILDING')
        if s(r.get('boq_id')).startswith('BOQ-') and s(r.get('building_id')).startswith('B')]

    # BOQ-006 ผูกกับอาคารในชีตทะเบียนโดยตรง (as-built ของ B006)
    linked = {(b['boq_id'], b['building_id']) for b in bld}
    for r in reg:
        direct = next((x for x in sheet(wb, '30_BOQ_REGISTER')
                       if s(x.get('boq_id')) == r['boq_id'] and s(x.get('building_id')).startswith('B')), None)
        if direct:
            key = (r['boq_id'], s(direct.get('building_id')))
            if key not in linked:
                bld.append({'boq_id': key[0], 'building_id': key[1], 'boq_budget': r['boq_value'],
                            'status': s(direct.get('สถานะ')), 'note': 'ผูกจากทะเบียน BOQ โดยตรง'})
                linked.add(key)
    return reg, bld

def build_pairs(wb):
    return [{
        'pair_id': s(r.get('pair_id')), 'label_a': s(r.get('อาคาร A')), 'label_b': s(r.get('อาคาร B')),
        'project_id': s(r.get('โครงการ')),
        'same_amount_count': int(num(r.get('จำนวนครั้งที่จ่ายพร้อมกันยอดเท่ากัน')) or 0),
        'status': s(r.get('สถานะ')), 'note': s(r.get('ใช้ทำอะไร')),
    } for r in sheet(wb, '09_DIM_BUILDING_PAIR', header_row=4) if s(r.get('pair_id')).startswith('P')]

THAI_MONTH = {'ม.ค.': 1, 'ก.พ.': 2, 'มี.ค.': 3, 'เม.ย.': 4, 'พ.ค.': 5, 'มิ.ย.': 6,
              'ก.ค.': 7, 'ส.ค.': 8, 'ก.ย.': 9, 'ต.ค.': 10, 'พ.ย.': 11, 'ธ.ค.': 12}

def build_funding(wb):
    out, n = [], 0
    for r in sheet(wb, '12_FACT_FUNDING_IN'):
        period, amount = s(r.get('period')), num(r.get('เงินทุนเข้าโครงการ (Funding In)'))
        if not period or not amount: continue
        mo = next((v for k, v in THAI_MONTH.items() if period.startswith(k)), None)
        if mo is None: continue
        n += 1
        out.append({
            'funding_id': f'FIN-{YEAR_CE % 100}{mo:02d}-{n:04d}',
            'funding_date': f'{YEAR_CE}-{mo:02d}-01', 'amount': amount,
            'company': 'RABBiZ Group', 'source': 'นำเข้าย้อนหลังจาก 12_FACT_FUNDING_IN',
            'accounting_status': 'เงินกู้ยืมกรรมการ', 'period_label': period,
            'value_source': 'นำเข้าย้อนหลัง',
        })
    return out

# ---------------------------------------------------------------- ฝั่งรายได้ (ข้อมูลอ้างอิง)
def build_rental_reference(wb):
    units = [{
        'unit_id': s(r.get('unit_id')), 'location': s(r.get('ทำเล')), 'location_code': s(r.get('รหัสทำเล')),
        'unit_label': s(r.get('ยูนิต')), 'official_name': s(r.get('รหัสอ้างอิง (ชื่ออาคารทางการ)')),
        'status': s(r.get('สถานะ')), 'tenant': s(r.get('ผู้เช่า')),
        'base_rent': num(r.get('ค่าเช่าฐาน/เดือน')), 'lessor': s(r.get('ผู้ให้เช่า (ออกบิลในนาม)')),
        'risk_level': s(r.get('ระดับความเสี่ยง')), 'building_id': s(r.get('building_id (จับคู่ v8)')) or None,
        'issue': s(r.get('ปัญหาที่พบ')),
    } for r in sheet(wb, '16_DIM_UNIT', header_row=4) if s(r.get('unit_id')).startswith('U')]

    lessors = [{
        'lessor_id': s(r.get('lessor_id')), 'lessor_name': s(r.get('ชื่อผู้ให้เช่า')),
        'entity_type': s(r.get('ประเภท')), 'tax_id': s(r.get('เลขทะเบียน/ผู้เสียภาษี')),
        'unit_count': int(num(r.get('จำนวนยูนิตที่ออกบิล')) or 0),
        'monthly_rent': num(r.get('ค่าเช่ารวม/เดือน')),
        'staff_count': int(num(r.get('จำนวนพนักงาน')) or 0),
        'monthly_payroll': num(r.get('ค่าจ้างรวม/เดือน')), 'note': s(r.get('หมายเหตุ')),
    } for r in sheet(wb, '17_DIM_LESSOR', header_row=4)
        if s(r.get('lessor_id')).startswith('LS') and s(r.get('lessor_id'))[2:].isdigit()]

    leases = [{
        'location_code': s(r.get('รหัส')), 'location': s(r.get('ทำเล')),
        'our_lessee': s(r.get('ผู้เช่าฝั่งเรา')), 'land_owner': s(r.get('เจ้าของที่ดิน')),
        'deed_no': s(r.get('เลขที่โฉนด')), 'area_wa': num(r.get('เนื้อที่ (ตร.ว.)')),
        'monthly_rent': num(r.get('ค่าเช่าที่ดิน/เดือน')),
        'start_date': s(r.get('วันเริ่ม')) or None, 'end_date': s(r.get('วันสิ้นสุด')) or None,
        'years_left': num(r.get('อายุคงเหลือ (ปี)')), 'renewal': s(r.get('สิทธิต่ออายุ')),
        'building_on_expiry': s(r.get('กรรมสิทธิ์อาคารเมื่อสิ้นสุด')),
    } for r in sheet(wb, '26_DIM_LAND_LEASE', header_row=4)
        if len(s(r.get('รหัส'))) == 3 and s(r.get('รหัส')).isupper()]

    pl = [{
        'location_code': s(r.get('รหัส')), 'location': s(r.get('ทำเล')),
        'units': int(num(r.get('ยูนิต')) or 0), 'vacant': int(num(r.get('ว่าง')) or 0),
        'rent_in': num(r.get('ค่าเช่ารับ/เดือน')), 'land_rent_out': num(r.get('ค่าเช่าที่ดิน/เดือน')),
        'margin_month': num(r.get('ส่วนต่าง/เดือน')), 'margin_year': num(r.get('ส่วนต่าง/ปี')),
        'years_left': num(r.get('อายุสัญญาคงเหลือ (ปี)')),
        'construction_spend': num(r.get('ก่อสร้าง มี.ค.-ส.ค. 2569')),
        'depreciation_year': num(r.get('ค่าเสื่อมต่อปีตามอายุสัญญา')), 'flag': s(r.get('ธงเตือน')),
    } for r in sheet(wb, '27_VIEW_LOCATION_PL', header_row=4)
        if len(s(r.get('รหัส'))) == 3 and s(r.get('รหัส')).isupper()]

    return units, lessors, leases, pl

# ---------------------------------------------------------------- รายการเบิก
def build_txn(wb, valid_buildings, valid_projects, valid_costcodes, staff_by_name):
    """1 แถวใน 10_FACT_TXN = 1 ใบเบิกที่ 'จ่ายแล้ว' + รายการย่อยแยกตาม cost_type"""
    # 25_VIEW_PAY_TO_STAFF — ใบที่จ่ายให้ทีมงานของเราเอง (พบหลังได้ชื่อจริงใน v9)
    to_staff = {}
    for r in sheet(wb, '25_VIEW_PAY_TO_STAFF', header_row=4):
        tid = s(r.get('txn_id'))
        if not tid.startswith('T'): continue
        to_staff[tid] = {
            'staff_user_id': s(r.get('รหัสผู้รับ')),
            'staff_name': s(r.get('ผู้รับเงิน')),
            'self_paid': 1 if s(r.get('เบิกเองจ่ายตัวเอง')) else 0,
        }

    extra_buildings, reqs, lines, seq = {}, [], [], {}
    for r in sheet(wb, '10_FACT_TXN'):
        tid = s(r.get('txn_id'))
        if not tid.startswith('T'): continue
        mo, dy = int(num(r.get('month')) or 0), int(num(r.get('day')) or 1)
        if not (1 <= mo <= 12): continue
        dy = min(max(dy, 1), 28 if mo == 2 else 30 if mo in (4, 6, 9, 11) else 31)
        d = f'{YEAR_CE}-{mo:02d}-{dy:02d}'
        proj = s(r.get('project_id')) or 'NA'
        if proj not in valid_projects: proj = 'NA'
        bldg = s(r.get('building_id'))
        if bldg not in valid_buildings:
            bldg = f'BX{proj}'
            extra_buildings.setdefault(bldg, {
                'building_id': bldg, 'project_id': proj, 'building_name': 'ไม่ระบุอาคาร',
                'design_code': None, 'work_nature': 'สร้างใหม่', 'status': 'กำลังทำ',
                'area_sqm': None, 'floors': None, 'is_building': 'N', 'budget': None,
                'value_source': 'นำเข้าย้อนหลัง',
                'note': 'อาคารพักสำหรับรายการนำเข้าที่ไม่มีรหัสอาคาร — ต้องย้ายเข้าอาคารจริง',
            })
        code = s(r.get('cost_code_v8')) or s(r.get('cost_code')) or 'OTH'
        if code not in valid_costcodes: code = 'OTH'
        amount = num(r.get('amount')) or 0.0
        # v9 แยกระดับ C2 ออกจาก D: C2 = ประมาณจากลักษณะงาน · D = ยังแยกไม่ได้จริง ๆ
        conf_raw = s(r.get('ระดับความเชื่อถือ'))
        conf = 'C2' if conf_raw.startswith('C2') else (conf_raw[:1] or 'D')
        payee = s(r.get('payee_name_raw'))
        staff = to_staff.get(tid) or (
            {'staff_user_id': staff_by_name.get(payee, ''), 'staff_name': payee, 'self_paid': 0}
            if payee in staff_by_name else None)

        key = f'{YEAR_CE % 100}{mo:02d}'
        seq[key] = seq.get(key, 0) + 1
        rid = f'REQ-{key}-{seq[key]:04d}'
        reqs.append({
            'request_id': rid, 'legacy_txn_id': tid, 'request_date': d,
            'requester_id': s(r.get('person_id')) or 'K',
            'project_id': proj, 'building_id': bldg,
            'vendor_id': s(r.get('vendor_id')) or None, 'payee_name_raw': payee,
            'has_vat': 'ไม่มี', 'vat_mode': 'แยก VAT',
            'amount_before_vat': amount, 'vat_amount': 0.0, 'total_amount': amount,
            'wht_amount': 0.0, 'net_amount': amount,
            'status': 'จ่ายแล้ว', 'confidence': conf, 'value_source': 'นำเข้าย้อนหลัง',
            'paid_to_staff': 1 if staff else 0,
            'staff_user_id': (staff or {}).get('staff_user_id') or None,
            'self_paid': (staff or {}).get('self_paid', 0),
            'note': s(r.get('building_name_raw')),
        })
        splits = [('ของ', num(r.get('cost_type_ของ')) or 0.0),
                  ('แรง', num(r.get('cost_type_แรง')) or 0.0),
                  ('เช่า', num(r.get('cost_type_เช่า')) or 0.0),
                  ('โสหุ้ย', num(r.get('cost_type_โสหุ้ย')) or 0.0)]
        used = round(sum(a for _, a in splits), 2)
        rest = round(amount - used, 2)
        for ct, amt in splits:
            if amt <= 0: continue
            lines.append({'request_id': rid, 'cost_code': code, 'cost_type': ct,
                          'description': s(r.get('building_name_raw')), 'qty': 1, 'unit': '',
                          'unit_price': amt, 'line_amount': amt, 'confidence': conf})
        if abs(rest) >= 0.01 or not any(a > 0 for _, a in splits):
            lines.append({'request_id': rid, 'cost_code': code, 'cost_type': 'ไม่ระบุ',
                          'description': s(r.get('building_name_raw')), 'qty': 1, 'unit': '',
                          'unit_price': rest, 'line_amount': rest, 'confidence': conf})
    return reqs, lines, list(extra_buildings.values())

# ---------------------------------------------------------------- main
def main():
    if not os.path.exists(SRC):
        sys.exit(f'ไม่พบไฟล์ต้นทาง: {SRC}')
    os.makedirs(OUT, exist_ok=True)
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    print(f'อ่าน {os.path.relpath(SRC, ROOT)}\n')

    users = build_users(wb);                write('users', users)
    projects = build_projects(wb);          write('projects', projects)
    project_ids = {p['project_id'] for p in projects}
    write('user_projects', build_user_projects(wb, project_ids))
    write('designs', build_designs(wb))
    write('cost_curve', build_cost_curve(wb))
    buildings = build_buildings(wb)
    write('building_aliases', build_building_aliases(wb))
    cost_codes = build_cost_codes(wb);      write('cost_codes', cost_codes)
    write('employees', build_employees(wb))
    vendors = build_vendors(wb);            write('vendors', vendors)
    write('items', build_items(wb))
    write('item_prices', build_item_prices(wb))
    write('rates', build_rates(wb))
    write('building_pairs', build_pairs(wb))
    write('funding_in', build_funding(wb))
    reg, bqb = build_boq(wb); write('boq_register', reg); write('boq_buildings', bqb)

    units, lessors, leases, pl = build_rental_reference(wb)
    write('rental_units', units); write('lessors', lessors)
    write('land_leases', leases); write('location_pl', pl)

    staff_by_name = {v['vendor_name']: v['staff_user_id'] for v in vendors if v['is_own_staff']}
    reqs, lines, extra = build_txn(
        wb, {b['building_id'] for b in buildings}, project_ids,
        {c['cost_code'] for c in cost_codes}, staff_by_name)
    write('buildings', buildings + extra)
    write('legacy_requests', reqs)
    write('legacy_request_lines', lines)

    total = round(sum(r['total_amount'] for r in reqs), 2)
    asset = {p['project_id'] for p in projects if p['is_group_asset']}
    asset_total = round(sum(r['total_amount'] for r in reqs if r['project_id'] in asset), 2)
    staff_total = round(sum(r['total_amount'] for r in reqs if r['paid_to_staff']), 2)
    print(f'\nใบเบิกนำเข้า {len(reqs):,} ใบ · {total:,.2f} บาท')
    print(f'  แยกเป็นทรัพย์สินกลุ่ม {asset_total:,.2f} · งานอื่น {total - asset_total:,.2f} บาท')
    print(f'  จ่ายให้ทีมงานของเราเอง {sum(1 for r in reqs if r["paid_to_staff"]):,} ใบ · {staff_total:,.2f} บาท')

if __name__ == '__main__':
    main()
